import openpyxl as xl
import random


def text_to_boolean(text):
    if "Y" in text or "y" in text or "M" in text or "m" in text:
        return True
    else:
        return False


def generate_lion_members(availability_spreadsheet):
    # Load Workbooks
    availability_book = xl.load_workbook(filename=availability_spreadsheet)

    # Get Proper Pages
    availability_page = availability_book["Availability"]

    # Set of members
    lion_set = set()

    # Create Members and add to member_set
    for r in range(2, availability_page.max_row + 1):
        member_info = availability_page[r]
        if text_to_boolean(member_info[1].value):
            lion_set.add(member_info[0].value)

    return lion_set


def generate_drum_members(availability_spreadsheet):
    # Load Workbooks
    availability_book = xl.load_workbook(filename=availability_spreadsheet)

    # Get Proper Pages
    availability_page = availability_book["Availability"]

    # Set of members
    drum_set = set()

    # Create Members and add to member_set
    for r in range(2, availability_page.max_row + 1):
        member_info = availability_page[r]
        if text_to_boolean(member_info[2].value):
            drum_set.add(member_info[0].value)

    return drum_set


def generate_box_members(availability_spreadsheet):
    # Load Workbooks
    availability_book = xl.load_workbook(filename=availability_spreadsheet)

    # Get Proper Pages
    availability_page = availability_book["Availability"]

    # Set of members
    box_set = set()

    # Create Members and add to member_set
    for r in range(2, availability_page.max_row + 1):
        member_info = availability_page[r]
        if text_to_boolean(member_info[3].value):
            box_set.add(member_info[0].value)

    return box_set


def generate_helpers(availability_spreadsheet):
    # Load Workbooks
    availability_book = xl.load_workbook(filename=availability_spreadsheet)

    # Get Proper Pages
    availability_page = availability_book["Availability"]

    # Set of members
    helper_set = set()

    # Create Members and add to member_set
    for r in range(2, availability_page.max_row + 1):
        member_info = availability_page[r]
        if text_to_boolean(member_info[5].value):
            helper_set.add(member_info[0].value)

    return helper_set


def check_used(row, name, rotation_page):
    for c in range(2, 14):
        if rotation_page.cell(row=row, column=c).value is not None \
                and name in rotation_page.cell(row=row, column=c).value:
            return True
        elif name == "Angie/Small" and "S" in rotation_page.cell(row=row, column=1).value:
            return True
        elif rotation_page.cell(row=row, column=c).value is not None \
                and "Lions" in rotation_page.cell(row=row, column=c).value \
                and "Lions" in name:
            return True
        elif rotation_page.cell(row=row, column=c).value is not None \
                and "Seniors" in rotation_page.cell(row=row, column=c).value\
                and "Seniors" in name:
            return True
    return False


def check_used_helper(row, name, rotation_page):
    for c in range(2, 14):
        if rotation_page.cell(row=row, column=c).value is not None \
                and name in rotation_page.cell(row=row, column=c).value:
            return True
        elif name == "Angie/Small" and "S" in rotation_page.cell(row=row, column=1).value:
            return True
    return False


# Load Workbooks
equipment_rotation_book = xl.load_workbook(filename="Equipment Rotation.xlsx")
availability_book = xl.load_workbook(filename="Equipment Rotation Availability.xlsx")
equipment_rotation_page = equipment_rotation_book["Nov 23"]
availability_page = availability_book["Availability"]

# Dictionary of Use
usage = dict()
usage["-"] = 0
for cell in availability_page["A"]:
    usage[cell.value] = 0

row = 2
while equipment_rotation_page.cell(row=row, column=1).value is not None:
    for col in range(2, 15):
        if equipment_rotation_page.cell(row=row, column=col).value is not None:
            usage[equipment_rotation_page.cell(row=row, column=col).value] = \
                usage[equipment_rotation_page.cell(row=row, column=col).value] + 1

    row = row + 1

row = 2
drum_set = generate_drum_members("Equipment Rotation Availability.xlsx")
box_set = generate_box_members("Equipment Rotation Availability.xlsx")
lion_set = generate_lion_members("Equipment Rotation Availability.xlsx")
helper_set = generate_helpers("Equipment Rotation Availability.xlsx")

while equipment_rotation_page.cell(row=row, column=1).value is not None:
    for col in range(2, 15):
        if equipment_rotation_page.cell(row=row, column=col).value is not None:
            continue

        member = "^"
        assigned_helper = False
        equipment_type = equipment_rotation_page.cell(row=1, column=col).value

        if "Drum" in equipment_type:
            member = random.choice(list(drum_set))
            while check_used(row, member, equipment_rotation_page):
                member = random.choice(list(drum_set))
        elif "Box" in equipment_type:
            member = random.choice(list(box_set))
            while check_used(row, member, equipment_rotation_page):
                member = random.choice(list(box_set))
        elif "Helper" in equipment_type:
            member = random.choice(list(helper_set))
            while check_used_helper(row, member, equipment_rotation_page):
                member = random.choice(list(helper_set))
            assigned_helper = True
        elif "Pole" in equipment_type:
            member = random.choice(list(helper_set))
            while check_used(row, member, equipment_rotation_page):
                member = random.choice(list(helper_set))
        else:
            member = random.choice(list(lion_set))
            while check_used(row, member, equipment_rotation_page):
                member = random.choice(list(lion_set))

        if member != "^":
            equipment_rotation_page.cell(row=row, column=col, value=member)
            usage[member] = usage[member] + 1
            if usage[member] > 4:
                drum_set.discard(member)
                box_set.discard(member)
                lion_set.discard(member)
                helper_set.discard(member)
        equipment_rotation_book.save('Equipment Rotation.xlsx')

    row = row + 1
